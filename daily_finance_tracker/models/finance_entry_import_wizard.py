from odoo import api, fields, models, _
import base64
import io
from datetime import datetime
import openpyxl
from odoo.exceptions import UserError

class FinanceEntryImportWizard(models.TransientModel):
    _name = 'finance.entry.import.wizard'
    _description = 'Import Past Finance Records'

    upload_file = fields.Binary(string='Excel File (.xlsx)', required=True)
    filename = fields.Char()
    sheet_name = fields.Char(string="Sheet Name", required=True)
    available_sheets = fields.Text(string="Available Sheets", readonly=True)
    
    @api.onchange('upload_file')
    def _onchange_upload_file(self):
        """Update available sheets when file is uploaded"""
        if self.upload_file:
            try:
                file_content = base64.b64decode(self.upload_file)
                workbook = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
                sheet_names = [s.strip() for s in workbook.sheetnames]
                
                # Reset sheet selection
                self.sheet_name = sheet_names[0] if sheet_names else ''
                self.available_sheets = ', '.join(sheet_names)
                
                return {
                    'warning': {
                        'title': _('Available Sheets'),
                        'message': _('Available sheets: %s') % ', '.join(sheet_names)
                    }
                }
            except Exception as e:
                self.available_sheets = ''
                self.sheet_name = ''
                raise UserError(_('Error reading Excel file: %s') % str(e))
        else:
            self.sheet_name = ''
            self.available_sheets = ''

    def _map_entry_type(self, excel_type):
        """Map Excel entry types to valid field values"""
        type_mapping = {
            'inflow': 'Inflow',
            'outflow': 'Outflow',
        }
        
        # Get the mapped value, or return the original if not found
        mapped_type = type_mapping.get(str(excel_type).strip(), str(excel_type).lower())
        
        # Validate against actual selection values
        finance_entry_model = self.env['finance.entry']
        type_field = finance_entry_model._fields.get('type')
        
        if hasattr(type_field, 'selection'):
            valid_types = [item[0] for item in type_field.selection]
            if mapped_type not in valid_types:
                for valid_type in valid_types:
                    if valid_type in mapped_type or mapped_type in valid_type:
                        return valid_type
        
        return mapped_type

    def action_import(self):
        if not self.upload_file:
            raise UserError(_("Please upload a valid Excel file."))
        
        if not self.sheet_name:
            raise UserError(_("Please enter a sheet name."))

        try:
            file_content = base64.b64decode(self.upload_file)
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
        except Exception as e:
            raise UserError(_("Failed to open Excel file. Please make usre it's a valid .xlsx file. \n\nDetails: %s") % str(e))
            
        # Match sheet name ignoring trailing spaces
        normalized_sheets = [s.strip() for s in workbook.sheetnames]
        if self.sheet_name.strip() not in normalized_sheets:
            available_sheets = ', '.join(workbook.sheetnames)
            raise UserError(_(
                "The sheet '%s' was not found in the uploaded Excel file.\n\n"
                " Available sheets: %s"
            ) % (self.sheet_name, ', '.join(workbook.sheetnames)))
        
        actual_sheet_name = workbook.sheetnames[normalized_sheets.index(self.sheet_name.strip())]
        sheet = workbook[actual_sheet_name]
       
        # Validate headers (assumes first now contains headers)
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
        expected_headers = ['Date', 'Type', 'Description', 'Amount']

        if len(header_row) < 4:
            raise UserError(_(
                "The sheet '%s' appears to have fewer than 4 columns.\n"
                "Expected at least the following columns: %s\n"
                "Found: %s"
            ) % (actual_sheet_name, ", ".join(expected_headers), ", ".join(str(h or "") for h in header_row)))

        for i, expected in enumerate(expected_headers):
            if str(header_row[i]).strip().lower() != expected.lower():
                raise UserError(_(
                    "Column %d in sheet '%s' should be '%s', but found '%s'.\n\n"
                    "Expected column order: %s"
                ) % (i+1, actual_sheet_name, expected, header_row[i] or "", ", ".join(expected_headers)))
        
        imported_count = 0
        skipped_count = 0
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                skipped_count += 1
                continue
            
            if len(row) < 4:
                skipped_count += 1
                continue
                
            date, entry_type, description, amount = row[:4]
            
            if not date or not entry_type or not description or amount is None:
                skipped_count += 1
                continue
            
            try:
                if isinstance(date, str):
                    for fmt in ('%Y-%m-%d','%m/%d/%Y', '%d/%m/%Y'):
                        try:
                            date = datetime.strptime(date, fmt).date()
                            break
                        except ValueError:
                            continue
                    else:
                        skipped_count += 1
                        continue

                mapped_type = self._map_entry_type(entry_type)
                amount_float = float(amount)               

                self.env['finance.entry'].create({
                    'date': date,
                    'description': str(description),
                    'amount': amount_float,
                    'type': mapped_type,
                    'user_id': self.env.uid,
                })
                imported_count += 1
                
            except Exception as e:
                skipped_count += 1
                continue
        
        workbook.close()
        
        # Summary message
        # return {
        #     'type': 'ir.actions.client',
        #     'tag': 'display_notification',
        #     'params': {
        #         'title': _('Import Complete'),
        #         'message': _(
        #             '%d finance record(s) successfully imported from sheet "%s".\n'
        #             '%d row(s) were skipped due to incomplete or invalid data.'
        #         ) % (imported_count, actual_sheet_name, skipped_count),
        #         'type': 'success',
        #         'sticky': False,
        #     }
        # }

        # Summary message
        return {
            'type': 'ir.actions.act_window',
            'name': _('Imported Finance Entries'),
            'res_model': 'finance.entry',
            'view_mode': 'tree,form',
            'domain': [('user_id', '=', self.env.uid)],
            'target': 'current',
        }
